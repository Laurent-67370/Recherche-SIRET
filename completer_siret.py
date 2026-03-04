"""
=============================================================
ABRAPA — Complétion et Audit SIRET fournisseurs
Utilise l'API SIRENE officielle (recherche-entreprises.api.gouv.fr)
=============================================================
MODES :

  complete (défaut) — Recherche les SIRET manquants :
    python completer_siret.py base_fournisseurs.xlsx

  verify  — Vérifie les SIRET existants contre SIRENE :
    python completer_siret.py base_fournisseurs.xlsx --mode verify

Options communes :
  --col-nom    Colonne nom fournisseur   (défaut : auto-détection)
  --col-siret  Colonne SIRET             (défaut : auto-détection)
  --col-cp     Colonne code postal       (défaut : auto-détection)
  --col-id     Colonne référence/ID Sage (défaut : auto-détection)
  --delay      Délai entre requêtes (s)  (défaut : 0.4)
  --dry-run    Analyse sans modifier le fichier
"""

import sys, time, re, argparse, unicodedata, requests
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

API_URL = "https://recherche-entreprises.api.gouv.fr/search"
SIRET_ZERO_RE = re.compile(r'^0+$')

# ── Couleurs openpyxl ─────────────────────────────────────
FILL_OK      = PatternFill("solid", fgColor="1A4731")
FILL_FOUND   = PatternFill("solid", fgColor="1A4731")
FILL_KEPT    = PatternFill("solid", fgColor="1E3A5F")
FILL_FERME   = PatternFill("solid", fgColor="4B1C1C")
FILL_DIFF    = PatternFill("solid", fgColor="3D2E00")
FILL_NOTFND  = PatternFill("solid", fgColor="2A2A2A")
FILL_INVALID = PatternFill("solid", fgColor="3D1A00")
FILL_MISSING = PatternFill("solid", fgColor="4B1C1C")
FILL_HEADER  = PatternFill("solid", fgColor="1E293B")

FONT_HEADER  = Font(bold=True, color="F97316", size=10)
FONT_OK      = Font(color="22C55E", bold=True)
FONT_FOUND   = Font(color="22C55E", bold=True)
FONT_KEPT    = Font(color="38BDF8")
FONT_FERME   = Font(color="EF4444", bold=True)
FONT_DIFF    = Font(color="FACC15", bold=True)
FONT_NOTFND  = Font(color="94A3B8")
FONT_INVALID = Font(color="FB923C", bold=True)
FONT_MISSING = Font(color="EF4444")
FONT_SIRET   = Font(color="FB923C", bold=True, name="Courier New")

AUDIT_STATUTS = {
    'ok':          ('OK SIRENE',             'Aucune action requise'),
    'ferme':       ('Ferme',                 'Bloquer le tiers dans Sage FRP 1000'),
    'nom-diff':    ('Nom different',         'Mettre a jour CT_Intitule dans Sage'),
    'ferme-diff':  ('Ferme + Nom different', 'Bloquer + verifier successeur eventuel'),
    'introuvable': ('Introuvable SIRENE',    'Verifier manuellement (SIRET obsolete ?)'),
    'invalide':    ('Format invalide',       'Corriger CT_Siret dans Sage (14 chiffres requis)'),
    'sans-siret':  ('Sans SIRET',            'SIRET a saisir dans Sage'),
}


# ── Helpers ───────────────────────────────────────────────
def is_siret_empty(val):
    if val is None:
        return True
    s = str(val).strip().replace(' ', '').replace('.', '').replace(',', '')
    if not s or s == 'nan':
        return True
    if SIRET_ZERO_RE.match(s):
        return True
    return len(re.sub(r'\D', '', s)) < 9

def clean_siret(val):
    if val is None:
        return ''
    return re.sub(r'\D', '', str(val).strip())

def format_siret(s):
    s = str(s).replace(' ', '')
    if len(s) == 14:
        return f"{s[:3]} {s[3:6]} {s[6:9]} {s[9:]}"
    return s

def detect_column(df, keywords):
    for col in df.columns:
        for kw in keywords:
            if kw.lower() in col.lower():
                return col
    return None

def norm_name(s):
    s = str(s).upper()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'\b(SAS|SARL|SA|SNC|SASU|EURL|EI|SCI|GIE|EHPAD|ASSO|ASSOCIATION|FONDATION|GROUPE)\b', '', s)
    s = re.sub(r'[^A-Z0-9]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def names_similar(a, b, threshold=0.5):
    if not a or not b:
        return True
    na, nb = norm_name(a), norm_name(b)
    if not na or not nb:
        return True
    if na == nb or na in nb or nb in na:
        return True
    wa = set(w for w in na.split() if len(w) > 2)
    wb = set(w for w in nb.split() if len(w) > 2)
    if not wa or not wb:
        return True
    return len(wa & wb) / max(len(wa), len(wb)) >= threshold


# ── API : recherche par nom ───────────────────────────────
def search_sirene(nom, code_postal='', retries=2):
    params = {'q': nom, 'per_page': 5}
    if code_postal:
        params['code_postal'] = str(code_postal).strip()[:5]
    for attempt in range(retries + 1):
        try:
            r = requests.get(API_URL, params=params, timeout=10)
            if r.status_code == 429:
                print(f"    Rate limit — attente 2s…")
                time.sleep(2); continue
            r.raise_for_status()
            data = r.json()
            results = data.get('results', [])
            if not results:
                return None
            best = results[0]
            etab = (best.get('matching_etablissements') or [best.get('siege', {})])[0]
            siret = clean_siret(etab.get('siret', ''))
            actif = etab.get('etat_administratif') == 'A'
            commune = etab.get('libelle_commune', '')
            nom_api = best.get('nom_complet') or best.get('nom_raison_sociale', '')
            naf = best.get('activite_principale', '')
            naf_lib = best.get('libelle_activite_principale', '')
            siege = best.get('siege', {}) or {}
            return {
                'siret':          siret,
                'actif':          actif,
                'nom_api':        nom_api,
                'adresse':        etab.get('adresse') or siege.get('adresse', ''),
                'cp':             etab.get('code_postal') or siege.get('code_postal', ''),
                'commune':        etab.get('libelle_commune') or siege.get('libelle_commune', ''),
                'forme_juridique':best.get('nature_juridique_label', ''),
                'naf':            f"{naf} - {naf_lib}" if naf else '',
                'tva':            best.get('numero_tva_intra', ''),
                'telephone':      siege.get('telephone', ''),
                'email':          siege.get('email', ''),
                'site_web':       siege.get('site_internet', ''),
            }
        except requests.RequestException as e:
            if attempt == retries:
                print(f"    Erreur API ({e})"); return None
            time.sleep(1)


# ── API : vérification par SIRET ─────────────────────────
def verify_sirene(siret, retries=2):
    params = {'q': siret, 'per_page': 1}
    for attempt in range(retries + 1):
        try:
            r = requests.get(API_URL, params=params, timeout=10)
            if r.status_code == 429:
                time.sleep(2); continue
            r.raise_for_status()
            data = r.json()
            results = data.get('results', [])
            if not results:
                return None
            best = results[0]
            etab = (best.get('matching_etablissements') or [best.get('siege', {})])[0]
            siret_api = clean_siret(etab.get('siret', ''))
            actif = etab.get('etat_administratif') == 'A'
            commune = etab.get('libelle_commune', '')
            nom_api = best.get('nom_complet') or best.get('nom_raison_sociale', '')
            naf = best.get('activite_principale', '')
            naf_lib = best.get('libelle_activite_principale', '')
            siege = best.get('siege', {}) or {}
            return {
                'siret_api':      siret_api,
                'actif':          actif,
                'nom_api':        nom_api,
                'adresse':        etab.get('adresse') or siege.get('adresse', ''),
                'cp':             etab.get('code_postal') or siege.get('code_postal', ''),
                'commune':        etab.get('libelle_commune') or siege.get('libelle_commune', ''),
                'forme_juridique':best.get('nature_juridique_label', ''),
                'naf':            f"{naf} - {naf_lib}" if naf else '',
                'tva':            best.get('numero_tva_intra', ''),
                'telephone':      siege.get('telephone', ''),
                'email':          siege.get('email', ''),
                'site_web':       siege.get('site_internet', ''),
            }
        except requests.RequestException as e:
            if attempt == retries:
                return None
            time.sleep(1)


# ── MODE COMPLETE ─────────────────────────────────────────
def process_complete(filepath, col_nom, col_siret, col_cp, delay, dry_run):
    path = Path(filepath)
    if not path.exists():
        print(f"Fichier introuvable : {filepath}"); sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  ABRAPA - Completion SIRET fournisseurs")
    print(f"  Fichier : {path.name}")
    print(f"{'='*60}\n")

    df = pd.read_excel(filepath, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    print(f"OK {len(df)} lignes - Colonnes : {list(df.columns)}\n")

    col_nom   = col_nom   or detect_column(df, ['nom','name','fournisseur','raison','intitule','ct_intitule'])
    col_siret = col_siret or detect_column(df, ['siret','ct_siret','siren'])
    col_cp    = col_cp    or detect_column(df, ['postal','cp','code_post','zip'])

    if not col_nom:
        print("Colonne Nom introuvable. Utilisez --col-nom."); sys.exit(1)
    if not col_siret:
        print("Colonne SIRET non trouvee - une colonne sera creee.")

    if col_siret and col_siret in df.columns:
        df['_siret_original'] = df[col_siret].copy()
    else:
        col_siret = 'SIRET'; df[col_siret] = ''; df['_siret_original'] = ''

    df['_statut'] = ''; df['_nom_api'] = ''; df['_actif'] = ''
    df['_commune'] = ''; df['_naf'] = ''
    df['_adresse'] = ''; df['_cp_api'] = ''; df['_forme_juridique'] = ''
    df['_tva'] = ''; df['_telephone'] = ''; df['_email'] = ''; df['_site_web'] = ''

    needs_siret = df[df[col_siret].apply(is_siret_empty)].index.tolist()
    has_siret   = df[~df[col_siret].apply(is_siret_empty)].index.tolist()

    print(f"Analyse :")
    print(f"  OK {len(has_siret)} fournisseurs avec SIRET existant (conserves)")
    print(f"  >> {len(needs_siret)} fournisseurs avec SIRET manquant\n")

    for idx in has_siret:
        df.at[idx, '_statut'] = 'conserve'

    if dry_run:
        print("DRY-RUN - lignes a traiter :\n")
        for idx in needs_siret[:20]:
            nom = df.at[idx, col_nom]
            cp  = df.at[idx, col_cp] if col_cp else ''
            print(f"  [{idx+2:>4}] {nom[:50]:<52} CP: {cp}")
        if len(needs_siret) > 20:
            print(f"  ... et {len(needs_siret)-20} autres")
        return

    found = notfound = 0
    print("Demarrage recherche SIRET via API SIRENE...\n")

    for i, idx in enumerate(needs_siret):
        nom = str(df.at[idx, col_nom]).strip()
        cp  = str(df.at[idx, col_cp]).strip() if col_cp and col_cp in df.columns else ''
        pct = int((i / len(needs_siret)) * 100)
        bar = '#' * (pct // 5) + '.' * (20 - pct // 5)
        print(f"  [{pct:>3}%] {bar} {i+1}/{len(needs_siret)} - {nom[:45]}", end='', flush=True)

        result = search_sirene(nom, cp)
        time.sleep(delay)

        if result and result['siret'] and len(result['siret']) >= 9:
            df.at[idx, col_siret]         = result['siret']
            df.at[idx, '_statut']         = 'trouve'
            df.at[idx, '_nom_api']        = result['nom_api']
            df.at[idx, '_actif']          = 'OUI' if result['actif'] else 'NON'
            df.at[idx, '_adresse']        = result.get('adresse', '')
            df.at[idx, '_cp_api']         = result.get('cp', '')
            df.at[idx, '_commune']        = result.get('commune', '')
            df.at[idx, '_forme_juridique']= result.get('forme_juridique', '')
            df.at[idx, '_naf']            = result['naf']
            df.at[idx, '_tva']            = result.get('tva', '')
            df.at[idx, '_telephone']      = result.get('telephone', '')
            df.at[idx, '_email']          = result.get('email', '')
            df.at[idx, '_site_web']       = result.get('site_web', '')
            found += 1
            print(f" -> OK {result['siret']}")
        else:
            df.at[idx, '_statut'] = 'introuvable'
            notfound += 1
            print(f" -> INTROUVABLE")

    print(f"\n{'='*60}")
    print(f"  {found} SIRET trouves, {notfound} introuvables, {len(has_siret)} conserves")
    print(f"{'='*60}\n")

    out_path = path.parent / (path.stem + '_siret_enrichi.xlsx')
    _save_complete_excel(df, col_nom, col_siret, out_path)
    print(f"Fichier sauvegarde : {out_path}\n")

    intro = df[df['_statut'] == 'introuvable'][[col_nom]].copy()
    if len(intro):
        csv_path = path.parent / (path.stem + '_siret_introuvables.csv')
        intro.to_csv(csv_path, sep=';', index=False, encoding='iso-8859-1')
        print(f"Introuvables : {csv_path}\n")


def _save_complete_excel(df, col_nom, col_siret, out_path):
    export_df = df.drop(columns=['_siret_original'], errors='ignore')
    export_df = export_df.rename(columns={
        '_statut':          'STATUT_SIRET',
        '_nom_api':         'NOM_SIRENE',
        '_actif':           'ACTIF_SIRENE',
        '_adresse':         'CT_Adresse',
        '_cp_api':          'CT_CodePostal',
        '_commune':         'CT_Ville',
        '_forme_juridique': 'CT_NatureJuridique',
        '_naf':             'CODE_NAF',
        '_tva':             'CT_NumTVAIntracomm',
        '_telephone':       'CT_Telephone',
        '_email':           'CT_Email',
        '_site_web':        'CT_Site',
    })
    export_df.to_excel(out_path, index=False, engine='openpyxl')
    wb = load_workbook(out_path)
    ws = wb.active
    for cell in ws[1]:
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        ml = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 4, 40)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def ci(name):
        try: return headers.index(name) + 1
        except: return None
    siret_col = ci(col_siret); statut_col = ci('STATUT_SIRET')
    for row in range(2, ws.max_row + 1):
        st = ws.cell(row, statut_col).value if statut_col else ''
        if siret_col:
            sc = ws.cell(row, siret_col)
            if st == 'trouve':     sc.font = FONT_FOUND; sc.fill = FILL_FOUND
            elif st == 'conserve': sc.font = FONT_KEPT;  sc.fill = FILL_KEPT
            elif st == 'introuvable': sc.font = FONT_MISSING; sc.fill = FILL_MISSING
        if statut_col:
            stc = ws.cell(row, statut_col)
            if st == 'trouve':        stc.value = 'Trouve';      stc.font = FONT_FOUND
            elif st == 'conserve':    stc.value = 'Conserve';    stc.font = FONT_KEPT
            elif st == 'introuvable': stc.value = 'Introuvable'; stc.font = FONT_MISSING
    ws.freeze_panes = 'A2'
    wb.save(out_path)


# ── MODE VERIFY ───────────────────────────────────────────
def process_verify(filepath, col_nom, col_siret, col_cp, col_id, delay, dry_run):
    path = Path(filepath)
    if not path.exists():
        print(f"Fichier introuvable : {filepath}"); sys.exit(1)

    print(f"\n{'='*62}")
    print(f"  ABRAPA - Audit SIRET fournisseurs (mode VERIFY)")
    print(f"  Fichier : {path.name}")
    print(f"{'='*62}\n")

    df = pd.read_excel(filepath, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    print(f"OK {len(df)} lignes - Colonnes : {list(df.columns)}\n")

    col_nom   = col_nom   or detect_column(df, ['nom','name','fournisseur','intitule','ct_intitule'])
    col_siret = col_siret or detect_column(df, ['siret','ct_siret'])
    col_cp    = col_cp    or detect_column(df, ['postal','cp','code_post','zip'])
    col_id    = col_id    or detect_column(df, ['ct_num','id','ref','code','numero'])

    if not col_siret:
        print("Colonne SIRET introuvable. Utilisez --col-siret."); sys.exit(1)

    print(f"Colonnes detectees :")
    print(f"  Nom    : {col_nom or '(absente)'}")
    print(f"  SIRET  : {col_siret}")
    print(f"  CP     : {col_cp or '(absente)'}")
    print(f"  ID/Ref : {col_id or '(absente)'}\n")

    df['_statut_audit']   = ''
    df['_nom_sirene']     = ''
    df['_actif_sirene']   = ''
    df['_adresse']        = ''
    df['_cp_api']         = ''
    df['_commune_sirene'] = ''
    df['_forme_juridique']= ''
    df['_naf']            = ''
    df['_tva']            = ''
    df['_telephone']      = ''
    df['_email']          = ''
    df['_site_web']       = ''
    df['_action']         = ''

    to_verify = []
    for idx, row in df.iterrows():
        siret = clean_siret(row.get(col_siret, ''))
        if not siret or re.fullmatch(r'0+', siret) or len(siret) < 9:
            df.at[idx, '_statut_audit'] = 'sans-siret'
            df.at[idx, '_action'] = AUDIT_STATUTS['sans-siret'][1]
        elif len(siret) != 14:
            df.at[idx, '_statut_audit'] = 'invalide'
            df.at[idx, '_action'] = AUDIT_STATUTS['invalide'][1]
        else:
            to_verify.append(idx)

    n_verify  = len(to_verify)
    n_invalid = (df['_statut_audit'] == 'invalide').sum()
    n_nosiret = (df['_statut_audit'] == 'sans-siret').sum()

    print(f"Analyse initiale :")
    print(f"  {n_verify} SIRET valides a verifier")
    print(f"  {n_invalid} SIRET avec format invalide")
    print(f"  {n_nosiret} sans SIRET\n")

    if dry_run:
        print("DRY-RUN - apercu :\n")
        for idx in to_verify[:20]:
            nom   = str(df.at[idx, col_nom]).strip() if col_nom else ''
            siret = clean_siret(df.at[idx, col_siret])
            print(f"  [{idx+2:>4}] {nom[:40]:<42} SIRET: {format_siret(siret)}")
        if len(to_verify) > 20:
            print(f"  ... et {len(to_verify)-20} autres")
        return

    cnt = {'ok':0, 'ferme':0, 'nom-diff':0, 'ferme-diff':0, 'introuvable':0}

    print(f"Demarrage audit SIRENE ({n_verify} fournisseurs)...\n")

    for i, idx in enumerate(to_verify):
        nom   = str(df.at[idx, col_nom]).strip() if col_nom else ''
        siret = clean_siret(df.at[idx, col_siret])
        pct   = int((i / n_verify) * 100)
        bar   = '#' * (pct // 5) + '.' * (20 - pct // 5)
        label = nom[:35] or siret
        print(f"  [{pct:>3}%] {bar} {i+1}/{n_verify} - {label}", end='', flush=True)

        result = verify_sirene(siret)
        time.sleep(delay)

        if result is None:
            df.at[idx, '_statut_audit'] = 'introuvable'
            df.at[idx, '_action']       = AUDIT_STATUTS['introuvable'][1]
            cnt['introuvable'] += 1
            print(" -> INTROUVABLE")
            continue

        actif   = result['actif']
        nom_api = result['nom_api']
        match   = names_similar(nom, nom_api) if nom else True

        df.at[idx, '_nom_sirene']      = nom_api
        df.at[idx, '_actif_sirene']    = 'OUI' if actif else 'NON'
        df.at[idx, '_adresse']         = result.get('adresse', '')
        df.at[idx, '_cp_api']          = result.get('cp', '')
        df.at[idx, '_commune_sirene']  = result.get('commune', '')
        df.at[idx, '_forme_juridique'] = result.get('forme_juridique', '')
        df.at[idx, '_naf']             = result['naf']
        df.at[idx, '_tva']             = result.get('tva', '')
        df.at[idx, '_telephone']       = result.get('telephone', '')
        df.at[idx, '_email']           = result.get('email', '')
        df.at[idx, '_site_web']        = result.get('site_web', '')

        if not actif and not match:   st = 'ferme-diff'
        elif not actif:               st = 'ferme'
        elif not match:               st = 'nom-diff'
        else:                         st = 'ok'

        df.at[idx, '_statut_audit'] = st
        df.at[idx, '_action']       = AUDIT_STATUTS[st][1]
        cnt[st] = cnt.get(st, 0) + 1

        icons = {'ok':'OK','ferme':'FERME','nom-diff':'NOM DIFF','ferme-diff':'FERME+NOM DIFF'}
        print(f" -> {icons.get(st,st)}" + (f" | {nom_api[:35]}" if st != 'ok' else ''))

    n_actions = sum(cnt[k] for k in cnt if k != 'ok')
    print(f"\n{'='*62}")
    print(f"  {cnt['ok']} OK (base a jour)")
    print(f"  {cnt['ferme']} etablissements fermes")
    print(f"  {cnt.get('nom-diff',0)} noms differents dans SIRENE")
    print(f"  {cnt.get('ferme-diff',0)} fermes + nom different")
    print(f"  {cnt['introuvable']} SIRET introuvables")
    print(f"  => {n_actions} actions requises dans Sage FRP 1000")
    print(f"{'='*62}\n")

    out_path = path.parent / (path.stem + '_audit_sirene.xlsx')
    csv_path = path.parent / (path.stem + '_audit_actions_sage.csv')

    _save_verify_excel(df, col_nom, col_siret, col_id, out_path)
    print(f"Rapport complet  : {out_path}")

    # CSV actions pour Sage FRP 1000 (priorité : fermés en premier)
    prio_map = {'ferme-diff':1, 'ferme':2, 'invalide':3, 'nom-diff':4, 'introuvable':5, 'sans-siret':6}
    actions_df = df[~df['_statut_audit'].isin(['ok', ''])].copy()
    actions_df['__prio'] = actions_df['_statut_audit'].map(prio_map).fillna(9)
    actions_df = actions_df.sort_values('__prio').drop(columns=['__prio'])

    export_cols = []
    if col_id and col_id in df.columns:     export_cols.append(col_id)
    if col_nom and col_nom in df.columns:   export_cols.append(col_nom)
    if col_siret and col_siret in df.columns: export_cols.append(col_siret)
    export_cols += ['_statut_audit', '_nom_sirene', '_actif_sirene', '_adresse', '_cp_api',
                   '_commune_sirene', '_forme_juridique', '_naf', '_tva',
                   '_telephone', '_email', '_site_web', '_action']
    actions_df[export_cols].to_csv(csv_path, sep=';', index=False,
                                    encoding='iso-8859-1', errors='replace')
    print(f"Actions Sage FRP : {csv_path}\n")


def _save_verify_excel(df, col_nom, col_siret, col_id, out_path):
    export_df = df.rename(columns={
        '_statut_audit':   'STATUT_AUDIT',
        '_nom_sirene':     'NOM_SIRENE',
        '_actif_sirene':   'ACTIF_SIRENE',
        '_adresse':        'CT_Adresse',
        '_cp_api':         'CT_CodePostal',
        '_commune_sirene': 'CT_Ville',
        '_forme_juridique':'CT_NatureJuridique',
        '_naf':            'CODE_NAF',
        '_tva':            'CT_NumTVAIntracomm',
        '_telephone':      'CT_Telephone',
        '_email':          'CT_Email',
        '_site_web':       'CT_Site',
        '_action':         'ACTION_SAGE',
    })
    export_df.to_excel(out_path, index=False, engine='openpyxl')

    wb = load_workbook(out_path)
    ws = wb.active

    for cell in ws[1]:
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_cells in ws.columns:
        ml = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 4, 45)

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def ci(name):
        try: return headers.index(name) + 1
        except: return None

    statut_col = ci('STATUT_AUDIT')
    siret_col  = ci(col_siret) if col_siret else None
    action_col = ci('ACTION_SAGE')

    STYLE_MAP = {
        'ok':         (FONT_OK,      FILL_OK),
        'ferme':      (FONT_FERME,   FILL_FERME),
        'nom-diff':   (FONT_DIFF,    FILL_DIFF),
        'ferme-diff': (FONT_FERME,   FILL_FERME),
        'introuvable':(FONT_NOTFND,  FILL_NOTFND),
        'invalide':   (FONT_INVALID, FILL_INVALID),
        'sans-siret': (FONT_NOTFND,  FILL_NOTFND),
    }

    for row in range(2, ws.max_row + 1):
        st = ws.cell(row, statut_col).value if statut_col else ''
        font, fill = STYLE_MAP.get(st, (FONT_NOTFND, FILL_NOTFND))
        if siret_col:
            ws.cell(row, siret_col).font = FONT_SIRET
        if statut_col:
            sc = ws.cell(row, statut_col)
            sc.value = AUDIT_STATUTS.get(st, ('',))[0]
            sc.font = font; sc.fill = fill
        if action_col:
            ws.cell(row, action_col).font = font

    ws.freeze_panes = 'A2'

    # Onglet recapitulatif
    ws_r = wb.create_sheet("Recapitulatif Audit")
    total = ws.max_row - 1
    ok_n    = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, statut_col).value == 'OK SIRENE') if statut_col else 0
    ferme_n = sum(1 for r in range(2, ws.max_row+1) if 'Ferme' in str(ws.cell(r, statut_col).value or '')) if statut_col else 0
    diff_n  = sum(1 for r in range(2, ws.max_row+1) if 'different' in str(ws.cell(r, statut_col).value or '')) if statut_col else 0
    nf_n    = sum(1 for r in range(2, ws.max_row+1) if 'Introuvable' in str(ws.cell(r, statut_col).value or '')) if statut_col else 0
    inv_n   = sum(1 for r in range(2, ws.max_row+1) if 'invalide' in str(ws.cell(r, statut_col).value or '')) if statut_col else 0
    pct     = round(ok_n/total*100, 1) if total else 0

    for r_data in [
        ["Statut SIRENE", "Nb fournisseurs", "% base"],
        ["OK SIRENE (base a jour)", ok_n, f"{pct}%"],
        ["Etablissements fermes", ferme_n, f"{round(ferme_n/total*100,1) if total else 0}%"],
        ["Noms differents", diff_n, f"{round(diff_n/total*100,1) if total else 0}%"],
        ["SIRET introuvables", nf_n, ""],
        ["Formats invalides", inv_n, ""],
        ["Total fournisseurs", total, "100%"],
    ]:
        ws_r.append(r_data)
    for cell in ws_r[1]:
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER

    wb.save(out_path)


# ── CLI ───────────────────────────────────────────────────
if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='ABRAPA - Completion et Audit SIRET fournisseurs',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python completer_siret.py export_sage.xlsx
  python completer_siret.py export_sage.xlsx --mode verify
  python completer_siret.py export_sage.xlsx --mode verify --dry-run
  python completer_siret.py base.xlsx --col-siret CT_Siret --col-id CT_Num --mode verify
        """
    )
    parser.add_argument('fichier',     help='Fichier Excel (.xlsx)')
    parser.add_argument('--mode',      choices=['complete','verify'], default='complete',
                        help='complete = cherche SIRET manquants | verify = audite SIRET existants')
    parser.add_argument('--col-nom',   default=None)
    parser.add_argument('--col-siret', default=None)
    parser.add_argument('--col-cp',    default=None)
    parser.add_argument('--col-id',    default=None, help='Colonne ID/ref Sage (CT_Num)')
    parser.add_argument('--delay',     type=float, default=0.4)
    parser.add_argument('--dry-run',   action='store_true')
    args = parser.parse_args()

    if args.mode == 'verify':
        process_verify(args.fichier, args.col_nom, args.col_siret, args.col_cp,
                       args.col_id, args.delay, args.dry_run)
    else:
        process_complete(args.fichier, args.col_nom, args.col_siret, args.col_cp,
                         args.delay, args.dry_run)
